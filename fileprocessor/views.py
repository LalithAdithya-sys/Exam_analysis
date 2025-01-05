from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from django.http import HttpResponse, FileResponse
from .forms import FileUploadForm
import openpyxl
import os

def upload_file(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            output_path = process_file(uploaded_file)
            request.session['output_file'] = output_path
            return render(request, 'fileprocessor/download.html')
    else:
        form = FileUploadForm()
    return render(request, 'fileprocessor/upload.html', {'form': form})


def process_file(uploaded_file):
    # Save the uploaded file
    input_path = f'temp/{uploaded_file.name}'
    with open(input_path, 'wb+') as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)

    # Process the file using your code
    wb_obj = openpyxl.load_workbook(input_path)
    questions, answers, Question_groups, student_names = [], [], [], []
    start_row, start_column = 1, 8

    sheet_obj = wb_obj[wb_obj.sheetnames[0]]
    for i in range(2, sheet_obj.max_row + 1):
        student_names.append(sheet_obj.cell(row=i, column=1).value)
    print(len(student_names))

    for i in range(1, sheet_obj.max_row + 1):
        for j in range(start_column, sheet_obj.max_column + 1, 2):
            if i == 1:
                questions.append(sheet_obj.cell(row=i, column=j).value)
            else:
                answers.append(sheet_obj.cell(row=i, column=j).value)

    sheet_obj = wb_obj[wb_obj.sheetnames[1]]
    for i in range(2, sheet_obj.max_row + 1):
        Question_groups.append(sheet_obj.cell(row=i, column=3).value)

    wb1_obj = openpyxl.Workbook()
    sheet_obj_1 = wb1_obj.active
    count = 0
    sheet_obj_1.cell(row=1, column=2).value = "Topic"
    for i in range(4, len(student_names) + 4):
        sheet_obj_1.cell(row=1, column=i).value = student_names[count]
        count += 1

    count = 0
    for j in range(4, len(student_names) + 4):
        for i in range(2, len(Question_groups) + 2):
            sheet_obj_1.cell(row=i, column=j).value = answers[count]
            count += 1

    count = 0
    for i in range(2, len(Question_groups) + 2):
        sheet_obj_1.cell(row=i, column=2).value = Question_groups[count]
        count += 1

    output_path = f'temp/processed_{uploaded_file.name}'
    wb1_obj.save(output_path)
    os.remove(input_path)  # Cleanup
    return output_path


def download_file(request):
    output_file = request.session.get('output_file')
    if output_file and os.path.exists(output_file):
        return FileResponse(open(output_file, 'rb'), as_attachment=True)
    return HttpResponse("No file to download.")